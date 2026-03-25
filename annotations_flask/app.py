# app.py
import csv
from io import StringIO
import webbrowser
from flask import Flask, make_response, render_template, request, redirect, send_file
from flask import jsonify
from openpyxl import Workbook
from utils.neuroglancer import (
    create_new_url_with_precomputed_annotations,
    create_new_url_with_multiple_precomputed_annotations,
    set_local_annotations,
)
import numpy as np

app = Flask(__name__)


@app.route("/", methods=["GET"])
def index():
    return redirect("/get_annotations")


@app.route("/get_annotations", methods=["GET", "POST"])
def get_annotations():
    new_url = None
    write_time = None
    if request.method == "POST":
        neuroglancer_url = request.values.get("neuroglancer_url")
        (
            annotation_type,
            all_annotations,
            write_time,
            new_url,
        ) = create_new_url_with_precomputed_annotations(neuroglancer_url)
        csv_data = StringIO()
        writer = csv.writer(csv_data)
        if annotation_type == "line":
            writer.writerow(
                [
                    "id",
                    "start x (nm)",
                    "start y (nm)",
                    "start z (nm)",
                    "end x (nm)",
                    "end y (nm)",
                    "end z (nm)",
                    "",
                    "neuroglancer url",
                ]
            )
        else:
            writer.writerow(
                [
                    "id",
                    "x (nm)",
                    "y (nm)",
                    "z (nm)",
                    "",
                    "neuroglancer url",
                ]
            )
        for idx in range(all_annotations.shape[0]):
            if idx == 0:
                writer.writerow([idx + 1, *all_annotations[idx, :], "", new_url])
            else:
                writer.writerow([idx + 1, *all_annotations[idx, :]])
        # return output
        return {
            "csv_data": csv_data.getvalue(),
            "new_url": new_url,
            "write_time": write_time,
        }
    return render_template("get_annotations.html")


@app.route("/get_multiple_annotations", methods=["GET", "POST"])
def get_multiple_annotations():
    if request.method == "POST":
        neuroglancer_url = request.values.get("neuroglancer_url")
        (
            annotation_layers,
            write_time,
            new_url,
        ) = create_new_url_with_multiple_precomputed_annotations(neuroglancer_url)

        # Build Excel workbook with one sheet per layer
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        layers_info = []
        seen_sheet_names = {}
        for layer_name, annotation_type, annotations in annotation_layers:
            # Excel sheet names max 31 chars
            sheet_name = layer_name[:31]
            # Deduplicate sheet names
            if sheet_name in seen_sheet_names:
                seen_sheet_names[sheet_name] += 1
                suffix = f" ({seen_sheet_names[sheet_name]})"
                sheet_name = sheet_name[: 31 - len(suffix)] + suffix
            else:
                seen_sheet_names[sheet_name] = 0

            ws = wb.create_sheet(title=sheet_name)

            if annotation_type == "line":
                ws.append(
                    [
                        "id",
                        "start x (nm)",
                        "start y (nm)",
                        "start z (nm)",
                        "end x (nm)",
                        "end y (nm)",
                        "end z (nm)",
                    ]
                )
            else:
                ws.append(["id", "x (nm)", "y (nm)", "z (nm)"])

            for idx in range(annotations.shape[0]):
                ws.append([idx + 1] + list(annotations[idx, :]))

            layers_info.append(
                {
                    "layer_name": layer_name,
                    "annotation_type": annotation_type,
                    "count": int(annotations.shape[0]),
                }
            )

        # Add a sheet with the neuroglancer URL (if annotations were saved)
        if new_url:
            url_ws = wb.create_sheet(title="neuroglancer url")
            url_ws.append(["neuroglancer url"])
            url_ws.append([new_url])

        # Save workbook to disk alongside the precomputed annotations
        excel_dir = "/groups/cellmap/cellmap/ackermand/neuroglancer_annotations/"
        excel_path = excel_dir + write_time + ".xlsx"
        wb.save(excel_path)

        return {
            "layers": layers_info,
            "new_url": new_url,
            "write_time": write_time,
            "excel_filename": write_time + ".xlsx",
        }
    return render_template("get_multiple_annotations.html")


@app.route("/download_excel/<filename>")
def download_excel(filename):
    excel_path = (
        "/groups/cellmap/cellmap/ackermand/neuroglancer_annotations/" + filename
    )
    return send_file(excel_path, as_attachment=True, download_name=filename)


@app.route("/set_annotations", methods=["GET", "POST"])
def set_annotations():
    return render_template("set_annotations.html")


@app.route("/get_editable_annotations", methods=["GET", "POST"])
def get_editable_annotations():
    new_url = None
    if request.method == "POST":
        neuroglancer_url = request.values.get("neuroglancer_url")
        new_url = set_local_annotations(neuroglancer_url)
        return {
            "new_url": new_url,
        }
    return render_template("get_editable_annotations.html")


# A function to add two numbers
@app.route("/add")
def add():
    a = request.args.get("a")
    b = request.args.get("b")
    return jsonify({"result": a + b})


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True)
